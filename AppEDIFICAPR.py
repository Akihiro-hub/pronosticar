import streamlit as st

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
import matplotlib.pyplot as plt
import numpy as np
import xgboost as xgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side  # 必要なモジュールをインポート
import bleach  # bleachをインポート

rubro = st.sidebar.selectbox("Herramientas de planificación a aplicar", ["Seleccione", "Plan de emprendimiento", "Plan de negocio en operación",  "Pronóstico de ventas", "Planificación de préstamos", "Plan de pagos de deuda e interés", "Plan del flujo de caja", "Plan de inversión", "Planificación de inventario",  "Análisis de punto de equilibrio", "Planificación de venta (Comedor)"])

if rubro == "Seleccione":
    st.write("## Aplicación digital para la elaboracion del plan de negocio :blue[(Proyecto EDIFICA)]")
    st.write("##### Esta aplicación contiene diferentes herramientas para facilitar la elaboración del plan de negocio. Dichas herramientas incluyen;") 
    st.write("(A) Plan de negocio en operación")
    st.write("(B) Plan de emprendimiento")
    st.write("(C) Plan de inversión")
    st.write("(D) Plan del flujo de caja")
    st.write("(E)Planificación de préstamos, etc")
    st.write(" :red[El dueño del negocio en operación puede aplicar (A) y los otros, dependiendo de la necesidad, mientras que el emprendedor deberá aplicar (B) (y los otros).]")

    st.write("###### (NOTA: Cada una de herramientas se presentará, dependiendo de su selección en las opciones presentadas a la izquierda.)")

elif rubro == "Pronóstico de ventas":
    # 過去12か月の売上データの初期値
    ventas_iniciales = [21000, 17500, 18000, 18500, 25000, 21000, 19000, 22000, 23500, 19500, 21000, 23000]
    # 過去12か月のその他の特徴量（Touristasは2022年の数値、千人単位、Cruceristas（通過者）は含まない。家族送金は2003～23年の月間平均。単位100万ドル）
    turistas = [44, 51, 71, 86, 69, 81, 85, 75, 54, 63, 71, 96]
    remesas = [477, 488, 581, 572, 618, 623, 606, 633, 599, 636, 573, 641]
    
    st.write("### :blue[Pronóstico (estimación) de ventas en próximos 12 meses]")
    st.write("###### (Herramienta de Inteligencia Artificial por Modelo XGBoost, ajustado del método de los mínimos cuadrados, para sectores de comercio y turísmo)")
    st.write("###### :red[Esta herramienta estima las ventas en futuro próximo, mediante la información sobre las ventas realizadas en estos 12 meses, los datos climáticos de la ciudad (a seleccionar) y el monto de remesas familiares por mes, el número de visitantes exteriores al país. Será probable que el resultado de estimación no sea precisa, debido a la limitación de los datos de variables explicativas.]")
    
    # 各都市のデータ
    ciudades = {
        "Tegucigalpa": {
            "lluvias": [0.4, 0.5, 0.5, 2.7, 9.8, 13.3, 10.6, 12.3, 15.0, 11.1, 3.7, 1.3],
            "temperaturas": [20, 21, 22, 24, 24, 23, 22, 23, 22, 22, 21, 20],
        },
        "San Marcos de Colón": {
            "lluvias": [0.2, 0.3, 0.8, 2.4, 8.9, 11.7, 8.5, 11.0, 13.5, 10.9, 3.5, 1.0],
            "temperaturas": [21, 22, 23, 24, 23, 22, 22, 22, 21, 21, 21, 21],
        },
        "Choluteca": {
            "lluvias": [0.1, 0.2, 0.7, 2.3, 8.9, 11.7, 8.6, 11.0, 13.6, 10.7, 3.2, 0.9],
            "temperaturas": [29, 29, 30, 31, 29, 28, 29, 29, 27, 27, 28, 29],
        },
        "Santa Rosa de Copán": {
            "lluvias": [1.9, 1.8, 1.6, 3.0, 8.2, 13.2, 12.3, 12.3, 13.4, 9.9, 4.8, 2.9],
            "temperaturas": [18, 19, 20, 22, 22, 22, 21, 22, 21, 20, 19, 18],
        },
        "San Pedro Sula": {
            "lluvias": [5.9, 4.7, 3.8, 3.4, 6.4, 11.1, 11.3, 11.2, 11.5, 10.3, 8.5, 7.2],
            "temperaturas": [23, 24, 25, 27, 28, 28, 27, 27, 27, 26, 24, 24],
        },
    }
    # 月の選択肢
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    
    st.write("##### :blue[Seleccione el mes actual y la ciudad cuyo clima es semejante al mismo de su lugar]")
    
    col1, col2 = st.columns(2)
    with col1:
        # 選択された月の初期値
        mes_actual = st.selectbox("Selecciona el mes actual", meses, index=7)
    
    with col2:
        # Select the city
        ciudad = st.selectbox("Selecciona la ciudad", list(ciudades.keys()))
    
    # Get the city's data
    lluvias = ciudades[ciudad]["lluvias"]
    temperaturas = ciudades[ciudad]["temperaturas"]
    
    # 月のインデックスを取得
    mes_index = meses.index(mes_actual)
    
    # ユーザーが売上データを入力
    st.write("##### :blue[Ingrese los datos de ventas de los últimos 12 meses]")
    
    # 各列に4か月分の売上データ入力フィールドを配置するための列の作成
    cols = st.columns(4)
    
    # 12か月前からの順序を保持し、各列に4か月分を表示
    for i in range(12):
        col_index = i // 3  # 0, 1, 2, 3 (4列)
        month_label = f"Hace {12 - i} meses ({meses[(mes_index - 12 + i) % 12]})"
        with cols[col_index]:
            ventas_iniciales[i] = st.number_input(month_label, value=ventas_iniciales[i], key=i)
    
    # データフレームの作成
    data = pd.DataFrame({
        'Ventas': ventas_iniciales,
        "Días de lluvias": lluvias[mes_index:] + lluvias[:mes_index],
        "Temperatura mínima del día": temperaturas[mes_index:] + temperaturas[:mes_index],
        'Visitantes exteriores al país': turistas[mes_index:] + turistas[:mes_index],
        "Remesas familiares": remesas[mes_index:] + remesas[:mes_index],
    })
    
    # 特徴量とターゲットの準備
    X = data[['Días de lluvias', 'Temperatura mínima del día', 'Visitantes exteriores al país', 'Remesas familiares']]
    y = data['Ventas']
    
    # データを訓練セットとテストセットに分割
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, shuffle=False)
    
    # XGBoostモデルの訓練
    model = xgb.XGBRegressor(objective='reg:squarederror', n_estimators=13)
    model.fit(X_train, y_train)
    
    # 12カ月先まで予測
    forecast_input = X.iloc[-1].values.reshape(1, -1)
    forecast = []
    for i in range(12):
        next_pred = model.predict(forecast_input)[0]
        forecast.append(next_pred)
        # 新しい特徴量の生成
        new_row = np.array([lluvias[(mes_index + i + 1) % 12], temperaturas[(mes_index + i + 1) % 12], turistas[(mes_index + i + 1) % 12], remesas[(mes_index + i + 1) % 12]]).reshape(1, -1)
        forecast_input = new_row
    
    forecast_df = pd.DataFrame(forecast, index=[f"{meses[(mes_index+i)%12]}" for i in range(12, 24)], columns=['Ventas'])
    forecast_df['Ventas'] = forecast_df['Ventas'].round(0).astype(int)  # 売上高を整数に丸める

    # 最小二乗法で傾きを計算
    from scipy.stats import linregress
    x = np.arange(len(ventas_iniciales))
    slope, intercept, _, _, _ = linregress(x, ventas_iniciales)

    # 傾きを加算して予測を修正
    forecast_df['Ventas'] = forecast_df['Ventas'] + slope * np.arange(1, 13)
    
    # 実績データと予測データの結合
    full_data = pd.concat([data, forecast_df])
    full_data.index = [f"Hace {12-i} meses ({meses[(mes_index-12+i)%12]})" for i in range(12)] + [meses[(mes_index+i)%12] for i in range(12, 24)]
    
    if st.button("Estimar (pronosticar) ventas futuras por la inteligencia artificial"):
    
        # グラフの表示
        st.subheader("Ventas realizadas y estimadas en los 24 meses")
        plt.figure(figsize=(12, 4))
        plt.plot(full_data.index[:12], full_data['Ventas'][:12], label='Ventas realizadas', color='blue', marker='o')
        plt.plot(full_data.index[12:], full_data['Ventas'][12:], label='Ventas estimadas', color='orange', marker='o')
        plt.xticks(rotation=45, ha='right')
        plt.legend(loc='upper left')
        plt.grid(True)
        plt.tight_layout()
        st.pyplot(plt)
    
        # 小数点以下を表示しない設定
        pd.options.display.float_format = '{:.0f}'.format   
        
        # 表の表示
        st.subheader("Datos de ventas realizadas y estimadas")
        st.write("Los datos de días de lluvia y otros indicadores no son exactamente del año pasado sino de los otros años de muestra.")
        resultados = pd.concat([data, forecast_df.round(0)])
        resultados.index = [f"Hace {12-i} meses ({meses[(mes_index-12+i)%12]})" for i in range(12)] + [meses[(mes_index+i)%12] for i in range(12, 24)]
        st.dataframe(resultados)
    
        # エクセルファイルのダウンロード
        st.subheader("Descargar Datos en Excel")
        def convert_df(df):
            return df.to_csv().encode('utf-8')
        csv = convert_df(resultados)
        st.download_button(label="Descargar datos en Excel como CSV", data=csv, file_name='prediccion_ventas.csv', mime='text/csv')
    
        
elif rubro == "Plan de negocio en operación":
    st.write("## :blue[Plan de negocio en operación]") 
    st.write("###### Esta herramienta facilita la planificación del monto a vender y el flujo de caja.") 
    
    def calculate_cash_flow(initial_cash, sales, material_cost, labor_cost, loan_repayment, other_fixed_costs, desired_profit):
        fixed_cost = labor_cost + loan_repayment + other_fixed_costs
        variable_ratio = material_cost / sales
        breakeven_sales = fixed_cost / (1 - variable_ratio)
        required_sales = (fixed_cost + desired_profit) / (1 - variable_ratio)
        
        cash_flow = {
            "Saldo del efecutivo al inicio": [],
            "Ingresos (Caja de entradas)": [],
            "Egresos (Caja de salidas)": [],
            "Saldo al final": []
        }
        for month in range(12):
            cash_inflow = sales
            cash_outflow = material_cost + labor_cost + loan_repayment + other_fixed_costs
            month_end_cash = initial_cash + cash_inflow - cash_outflow
            cash_flow["Saldo del efecutivo al inicio"].append(initial_cash)
            cash_flow["Ingresos (Caja de entradas)"].append(cash_inflow)
            cash_flow["Egresos (Caja de salidas)"].append(cash_outflow)
            cash_flow["Saldo al final"].append(month_end_cash)
            initial_cash = month_end_cash
        return breakeven_sales, required_sales, cash_flow, fixed_cost, variable_ratio

    def generate_excel(cash_flow):
        wb = Workbook()
        ws = wb.active
        ws.title = "Presupuesto del flujo de caja"

        headers = ["", "1r mes", "2do mes", "3r mes", "4to mes", "5to mes", "6to mes", "7mo mes", "8vo mes", "9no mes", "10mo mes", "11 mes", "12 mes"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, (key, values) in enumerate(cash_flow.items(), 2):
            ws.cell(row=row_num, column=1, value=key)
            for col_num, value in enumerate(values, 2):
                ws.cell(row=row_num, column=col_num, value=value)

        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        return excel_data

    col1, col2 = st.columns(2)
    with col1:
        sales = st.number_input("Monto estimado de venta mensual (¿Cuánto monto su negocio vende al mes en Lps?):", min_value=0, value=16000, step=1, format="%d")
        desired_profit = st.number_input("Meta de ganancias mensuales (¿Cuánto desea ganar al mes en Lps?):", min_value=0, value=5000, step=1, format="%d")
        initial_cash = st.number_input("Saldo inicial del efecutivo (¿Cuánto monto de efecutivo comercial tiene actualmente en Lps?):", min_value=0, value=4500, step=1, format="%d")
    with col2:
        material_cost = st.number_input("Costo mensual de materias primas (y otros costos variables, Lps):", min_value=0, value=6000, step=1, format="%d")
        labor_cost = st.number_input("Remuneraciones mensuales de trabajadores como costo fijo (Lps):", min_value=0, value=4000, step=1, format="%d")
        loan_repayment = st.number_input("Pago mensual de deuda (como costo fijo, Lps):", min_value=0, value=0, step=1, format="%d")
        other_fixed_costs = st.number_input("Otros costos fijos, tales como alquiler de la tienda, electricidad, etc (Lps):", min_value=0, value=4500, step=1, format="%d")
       
    if st.button("Elaborar el plan operativo de negocio (planificación de venta y flujo de caja)"):
        breakeven_sales, required_sales, cash_flow, fixed_cost, variable_ratio = calculate_cash_flow(
            initial_cash, sales, material_cost, labor_cost, loan_repayment, other_fixed_costs, desired_profit)

        months = ["1r mes", "2do mes", "3r mes", "4to mes", "5to mes", "6to mes", "7mo mes", "8vo mes", "9no mes", "10mo mes", "11 mes", "12 mes"]
        df = pd.DataFrame(cash_flow, index=months).T
        st.write("#### :blue[(1) Presupuesto del flujo de caja por 12 meses]") 
        st.dataframe(df)

        excel_data = generate_excel(cash_flow)
        st.download_button(
            label="Descargar la tabla EXCEL",
            data=excel_data,
            file_name="business_plan.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.write("###### Puede descargar la tabla en Excel. Es recomendable elaborar el plan del flujo de caja de manera más precisa, aplicando la otra herramienta, puesto que la tabla presentada arriba es de versión muy resumida.") 

        st.write("#### :blue[(2) Planificación de ventas, en base al análisis del punto de equilibrio]") 
        st.write(f"Ventas al mes en el punto de equilibrio: {breakeven_sales:.2f} Lps")
        st.write(f"Ventas necesarias para lograr la meta de ganancias al mes: {required_sales:.2f} Lps")

        fig, ax = plt.subplots()
        
        sales_range = list(range(int(breakeven_sales * 0.8), int(required_sales * 1.2), 100))
        total_costs = [fixed_cost + (variable_ratio * s) for s in sales_range]
        
        ax.plot(sales_range, total_costs, color='skyblue', label="Costos totales (Costos fijos + Costos variables)", marker='o')
        ax.plot(sales_range, sales_range, color='orange', label="Venta", marker='o')
        
        ax.set_title("Análisis de punto de equilibrio")
        ax.set_xlabel("Venta (Lps)")
        ax.set_ylabel("Costos y ventas (Lps)")
        
        ax.axvline(breakeven_sales, color='red', linestyle='--', label=f"Punto de equilibrio: {breakeven_sales:.2f} Lps")
        
        ax.fill_between(sales_range, total_costs, sales_range, where=[s > breakeven_sales for s in sales_range], color='skyblue', alpha=0.3, interpolate=True)
        
        mid_x = (required_sales + breakeven_sales) / 2
        mid_y = (max(total_costs) + max(sales_range)) / 2
        ax.text(mid_x, mid_y, "Ganancia = Área del color azul claro", color="blue", fontsize=7, ha="center")

        ax.legend()  # Show the legend
        st.pyplot(fig)

elif rubro == "Planificación de inventario":
    st.write("## :blue[Planificación de inventario de seguridad]") 
    st.write("###### Esta herramienta facilita la calculación del volumen de inventario de seguridad, que se refiere a la cantidad necesaria a mantener siempre para evitar escasez, en ciertas materias importantes.")  
    st.write("###### En el siguiente ejemplo se muestra un caso de maderas aserradas como la materia prima principal de la carpintería, mientras que esta herramienta es aplicable para otros negocios también.")
    st.write("###### Es importante calcular el volumen de inventario de seguridad, ya que el mismo se relaciona directamente al monto necesario del capital de trabajo.")
    col1, col2 = st.columns(2)
    with col1:
        a = st.number_input("¿Hace 5 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 30)
        b = st.number_input("¿Hace 4 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 25)
        c = st.number_input("¿Hace 3 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 45)
    with col2:
        d = st.number_input("¿Hace 2 días (o semana) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 37)
        e = st.number_input("¿Ayer (o semana pasada) cuántas piezas de madera aserrada se consumieron?", 0, 10000, 18)
        g = st.number_input("¿Cuánto días (o semanas) debe esperar la recepción de maderas después de la colocación de la orden?", 0, 300, 5)
    data = [a, b, c, d, e]
    SD = np.std(data, ddof=1) 
    import math
    Inventario_seguridad1 = 2.33 * SD * math.sqrt(g)
    Inventario_seguridad5 = 1.64 * SD * math.sqrt(g)   
    Inventario_seguridad10 = 1.28 * SD * math.sqrt(g)

    if st.button("Calcular el volumen de inventario de seguridad)"):
        st.write("##### Resultado de cálculo:") 
        col1, col2 = st.columns(2)
        with col1:
            st.write("##### :green[Volumen de inventario de seguridad]")
            st.write("###### Caso A: Inventario de seguridad con la probabilidad de escasez de 1% (piezas):")
            st.text(round(Inventario_seguridad1))
            st.write("###### Caso B: Inventario de seguridad con la probabilidad de escasez de 5% (piezas):")
            st.text(round(Inventario_seguridad5))
            st.write("###### Caso C: Inventario de seguridad con la probabilidad de escasez de 10% (piezas):")
            st.text(round(Inventario_seguridad10))  
        with col2:
            st.write("##### :green[Volumen al punto de ordenar ]")
            st.write("###### Volumen de inventario en posesión al punto de ordenar en Caso A (piezas):")
            st.text(round(Inventario_seguridad1+np.mean(data)*g))
            st.write("###### Volumen de inventario en posesión al punto de ordenar en Caso B (piezas):")
            st.text(round(Inventario_seguridad5+np.mean(data)*g))
            st.write("###### Volumen de inventario en posesión al punto de ordenar en Caso C (piezas):")
            st.text(round(Inventario_seguridad10+np.mean(data)*g))  
        st.write("###### :red[NOTA: Además del inventario de seguridad, la empresa también necesita tener cierto volumen del inventario para su consumo durante el período de espera después de colocación de la orden de materias primas, por lo que el volumen de inventario a tener al punto de ordenar debe ser mayor que el inventario de seguridad. En otras palabras, el volumen al punto de colocación de la orden puede ser; Promedio de consumos diarios x Días de espera + Inventario de seguridad.]")

elif rubro == "Planificación de préstamos":
    st.write("## :blue[Planificación de préstamos]") 
    st.write("###### El monto disponible para el préstamo dependerá de (i) cuota mensual a poder pagar, (ii) tasa de interés, y (iii) período de amortización, como se puede calcular mediante esta herramienta.")
    a = st.number_input("Cuota mensual (Lps)", 0, 1000000000, 12000)
    b = st.number_input("Tasa anual de interés %", 0, 100, 22)
    c = st.number_input("Periodo de amortización (meses)", 0, 100, 12)
    d = (a * ((1 + b/1200)**c - 1)) / (b/1200 * (1 + b/1200)**c)

    if st.button("Calcular"):
        st.write("##### :blue[Resultado del cálculo: Monto total disponible para el préstamo (Lps):]")
        st.text(round(d))

elif rubro == "Planificación de venta (Comedor)":
    st.write("## :blue[Planificación del monto de ventas en un comedor]") 
    st.write("###### El monto de la venta de un restaurante, comedor o cafetería se puede estimar, en base al número de asientos, aplicando esta calculadora.")  
    a = st.number_input("¿Cuánto asientos tiene el comedor?", 0, 1000, 20)
    b = st.number_input("Tasa de ocupación de los asientos por los clientes (%)", 0, 100, 50)
    c = st.number_input("Veces estimadas de rotación de los clientes al día", 1, 10, 3)
    d = st.number_input("Promedio estimado de la venta por cliente (Lps)", 1, 1000, 125)
    e = st.number_input("Días de operación al mes (Días)", 1, 31, 28)
    st.write("###### :red[La tasa de ocupación puede ser 50%, ya que sólo dos personas pueden ocupar la mesa para cuatro personas. La rotacion de los clientes al día puede ser 4 o 5 veces, como 2 rotaciones a horas de almuerzo y 2 rotaciones a horas de cena.]")
    
    E = a*d*(b/100)*c

    if st.button("Estimar el monto de ventas"):
        st.write("##### Resultado del cálculo: Monto esperado de la venta diaria")
        st.text(E)
        st.write("##### Resultado del cálculo: Monto esperado de la venta mensual")
        st.text(E*e)

elif rubro == "Análisis de punto de equilibrio":
    st.write("## :blue[Análisis de punto de equilibrio]") 
    st.write("###### Se puede calcular la meta de venta, en base al análisis del punto de equilibrio. Mientras que el siguiente ejemplo se refiere a un caso de panadería, esta calculadora se puede aplicar en cualquier negocio.")  
    a = st.number_input("Precio unitario (¿cuánto cuesta un paquete de panes a vender como promedio?, Lps)", 1, 100000000000, 30)
    b = st.number_input("Costo variable unitario (¿cuánto vale el costo de materias primas para un paquete?, Lps)", 0, 100000000000, 6)
    c = st.number_input("Costo fijo mensual (alquiler del espacio, depreciación de la maquina, costo de electricidad, etc., Lps)", 1, 100000000000, 6000)
    d = st.number_input("Ganancias mensuales que desea (Lps)", 1, 10000000000, 1500)
    CM = a-b
    CMR = CM/a

    if st.button("Calcular"):
        st.write("##### Monto de la venta necesaria para alcanzar la ganancia deseada (Lps)")
        st.text(round((c+d)/(CMR)))
        st.write("##### Punto de equilibrio en venta (Lps)")
        st.text(round(c/CMR))


elif rubro == "Plan de pagos de deuda e interés":
    st.write("## :blue[Plan de pagos de deuda e interés (Cálculo de amortización de préstamo)]") 
    st.write("###### Esta herramienta calcula el monto de la cuota mensual, la proporción de intereses y capital en un préstamo de amortización constante y genera el cuadro de amortización del préstamo.")  

    # 入力項目
    principal = st.number_input("Monto del préstamo (Lps):", min_value=0, value=70000, step=1000, format="%d")
    annual_rate = st.number_input("Tasa de interés anual (%):", min_value=0.0, value=26.0, step=0.1, format="%f")
    months = st.number_input("Plazo de reembolso (meses):", min_value=1, value=15, step=1, format="%d")

    # 計算を行うボタン
    if st.button("Calcular el cuadro de amortización"):
        # 月利の計算
        monthly_rate = annual_rate / 100 / 12

        # 毎月の返済額の計算
        monthly_payment = (principal * monthly_rate * (1 + monthly_rate) ** months) / ((1 + monthly_rate) ** months - 1)

        # 初期設定
        balance = principal
        schedule = []

        # 各月の償還表を作成
        for month in range(1, months + 1):
            interest_payment = balance * monthly_rate
            principal_payment = monthly_payment - interest_payment
            balance -= principal_payment
            schedule.append([month, round(monthly_payment), round(principal_payment), round(interest_payment), round(balance)])

        # データフレームに変換し、インデックスを表示しない
        df = pd.DataFrame(schedule, columns=["Mes", "Pago mensual (Lps)", "Pago a capital (Lps)", "Interés (Lps)", "Saldo restante (Lps)"])
        df = df.reset_index(drop=True)  # インデックスをリセットして削除

        # 結果の表示（インデックスをリセットして表示）
        st.write("#### Cuadro de Amortización en base al plan de cuotas niveladas")
        st.dataframe(df.reset_index(drop=True))

        # Excelファイルのダウンロードオプション
        def generate_excel(dataframe):
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                dataframe.to_excel(writer, index=False, sheet_name="Amortización")
            return output.getvalue()

        excel_data = generate_excel(df)
        st.download_button(
            label="Descargar el cuadro en Excel",
            data=excel_data,
            file_name="cuadro_de_amortizacion.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


elif rubro == "Plan de inversión":
    st.write("## :blue[Planificación de inversión]") 
    st.write("###### Esta herramienta calcula ciertos indicadores para poder evaluar la factibilidad económica del proyecto de inversión.")  

    a = st.number_input("¿Cuánto se debe invertir al inicio del proyecto (Lps)?", 0, 10000000000000, 200000)
    b = st.number_input("¿Cuál es tasa de costo del capital del negocio (%)?", 0, 100, 14)
    c = st.number_input("¿Cuánto podrá ganar al año por el proyecto de inversión? (De manera más precisa tiene que decirse como el flujo anual de caja, que es casi igual a ganancias menos depreciación: Lps)", 1, 1000000000000, 70000)
    d = st.number_input("Duración del proyecto (años)", 1, 100, 4)
        
    lst = [c for i in range(d)]
    lst0 = [-1 * a]
    lst = lst0 + lst
    npv = sum(lst / (1 + b/100) ** t for t, lst in enumerate(lst)) 
    rate = b/100

    payback = 1/rate - (1/(rate*(1+rate)**d))

    if st.button("Calcular los indicadores de la inversión"):
        st.write("##### Valor Presente Neto (VPN) de la inversión (Lps):")
        st.text(round(npv))
        st.write("##### Periodo máximo aceptable para recaudación del fondo invertido (meses):")
        st.text(round(payback*12))
        st.write("###### :red[Un proyecto con el VPN negativo o insuficiente se debe rechazar. El segundo indicador es para la referencia teórica, y el empresario deberá recuperar el fondo invertido lo antes posible.]") 

elif rubro == "Plan del flujo de caja":

    st.write("## :blue[Plan del flujo de caja]") 
    st.write("###### Es importante que el empresario o emprendedor elabore el presupuesto del flujo de caja para ver si el negocio puede mantener bien su liquidez o no.")  

    # Excel file path
    file_path = 'plan_de_flujo_de_caja.xlsx'

    # Read the Excel file
    df = pd.read_excel(file_path, sheet_name=None)

    # Display the dataframe as a non-interactive table
    for sheet_name, data in df.items():
        st.write(f"### {sheet_name}")
        st.dataframe(data)  # Displaying the dataframe

    # Download button
    st.write("Se presenta arriba un ejemplo del plan de flujo de caja. Puede descargar el formato Excel para elaborar su propio plan.")
    st.write(" :red[NOTA; Balance al final = Balance al inicio + Ingresos - Todos los egresos]")

    output = BytesIO()

    # Save the Excel file to the buffer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, data in df.items():
            data.to_excel(writer, index=False, sheet_name=sheet_name)

    output.seek(0)

    # Create the download link
    st.download_button(
        label="Descargar la tabla Excel",
        data=output,
        file_name="plan_de_flujo_de_caja.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

elif rubro == "Plan de emprendimiento":

    # エクセルファイルのパス
    file_path = 'Plan emprendimiento.xlsx'

    # エクセルファイルの読み込み
    sheets = pd.read_excel(file_path, sheet_name=None)

    # 各シートをインタラクティブに表示
    st.write("# :blue[Plan de emprendimiento]") 
    st.write("###### El plan de emprendimiento se puede elaborar, mediante los siguientes pasos (1) concretar las ideas del negocio,y (2) preparar el plan financiero al inicio del negocio.")  
    
    st.write("## :blue[Paso 1: Concretar las ideas sobre el negocio]") 
    st.write("Este primer paso se puede desarrollar, mediante la llenada de las siguientes dos tablas. La primera tabla facilita concretar las ideas del negocio a montar. La segunda apoya que tenga las ideas sobre compras y ventas en cuanto al negocio.")  
    edited_sheets = {}

    for sheet_name, df in sheets.items():
        st.subheader(f"Tabla: {sheet_name}")
        
        # 数字部分を文字列に変換
        df = df.astype(str)
        
        # データフレームを編集可能に表示
        edited_df = st.data_editor(df)
        
        # 文字列を元のデータ型に戻す
        edited_sheets[sheet_name] = edited_df.apply(pd.to_numeric, errors='ignore')

    # 編集されたデータフレームを保存するボタンを表示
    if st.button("Guardar las tablas en Excel"):
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, edited_df in edited_sheets.items():
                edited_df.to_excel(writer, index=False, sheet_name=sheet_name)
                
                # ワークシートを取得
                worksheet = writer.sheets[sheet_name]
                
                # 列幅を調整
                for col in worksheet.columns:
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[col_letter].width = 30  # すべての列幅を30に設定

                # 格子線を追加
                thin_border = Border(left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin'))

                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border

        output.seek(0)

        # ダウンロードリンクを作成
        st.download_button(label="Descargar", data=output, file_name="Plan_emprendimiento_editado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # 資金計画部分
    st.write("## :blue[Paso 2: Elaborar el plan financiero al inicio]") 
    st.write("Como el segundo paso, el emprendedor deberá elaborar el plan financiero al inicio del negocio, retroalimentando el paso anterior. El plan deberá identidicar el capital a necesitar y cómo adquirirlo.") 
    st.write("Con relación al capital de trabajo, será importante estimar el monto necesario para los primeros 3 meses de operación, considerando que las ventas podrán ser inestables al inicio del negocio.") 

    # 初期データ
    data1 = {
        "Asuntos": ["2 microondas", "3 fogones", "Remodelación de la tienda", "Otros"] + [""] * 1,
        "Monto (Lps.)": [8000, 21000, 90000, 11000] + [0] * 1 
    }

    data2 = {
        "Asuntos": ["Materias primas (harina de trigo, huevos, etc.)", "Agua y electricidad", "Otros"] + [""] * 2,
        "Monto (Lps.)": [30000, 12000, 8000] + [0] * 2
    }

    # データフレーム作成
    df1 = pd.DataFrame(data1)
    df2 = pd.DataFrame(data2)
    
    # 編集用のインタラクティブなテーブルを作成する関数
    def edit_table(df, column_name):
        edited_data = []
        for i, row in df.iterrows():
            asunto = bleach.clean(st.text_input(f"Asunto {i+1}", row['Asuntos'], key=f"asunto_{i}_{column_name}"))
            monto = st.number_input(f"Monto {i+1} (Lps.)", value=row['Monto (Lps.)'], min_value=0, step=100, key=f"monto_{i}_{column_name}")
            edited_data.append([asunto, monto])
        return pd.DataFrame(edited_data, columns=['Asuntos', 'Monto (Lps.)'])

    # Capital de inversión a necesitar
    st.subheader("Tabla: Capital de inversión a necesitar")

    # 5列表示を作成
    cols = st.columns(5)

    # 各列にAsuntoとMontoを順番に配置
    for i in range(5):
        with cols[i]:
            if i < len(df1):
                bleach.clean(st.text_input(f"Asunto {i+1}", df1.at[i, 'Asuntos'], key=f"asunto_{i}_inversion"))
                st.number_input(f"Monto {i+1} (Lps.)", value=df1.at[i, 'Monto (Lps.)'], min_value=0, step=100, key=f"monto_{i}_inversion")

    # 合計の計算と表示
    editable_df1 = pd.DataFrame({
        "Asuntos": [st.session_state[f"asunto_{i}_inversion"] for i in range(5)],
        "Monto (Lps.)": [st.session_state[f"monto_{i}_inversion"] for i in range(5)]
    })
    total1 = editable_df1["Monto (Lps.)"].sum()
    st.write(f"**Total Capital de Inversión:** {total1} Lps.")

    # Capital de trabajo a necesitar para primeros 3 meses de la operación
    st.subheader("Tabla: Capital de trabajo para primeros 3 meses")

    # 5列表示を作成（2つ目のテーブル）
    cols2 = st.columns(5)

    for i in range(5):
        with cols2[i]:
            if i < len(df2):
                bleach.clean(st.text_input(f"Asunto {i+1}", df2.at[i, 'Asuntos'], key=f"asunto_{i}_trabajo"))
                st.number_input(f"Monto {i+1} (Lps.)", value=df2.at[i, 'Monto (Lps.)'], min_value=0, step=100, key=f"monto_{i}_trabajo")

    # 合計の計算と表示
    editable_df2 = pd.DataFrame({
        "Asuntos": [st.session_state[f"asunto_{i}_trabajo"] for i in range(5)],
        "Monto (Lps.)": [st.session_state[f"monto_{i}_trabajo"] for i in range(5)]
    })
    total2 = editable_df2["Monto (Lps.)"].sum()
    st.write(f"**Total Capital de Trabajo:** {total2} Lps.")

    # Tabla 3: Fuentes del capital necesario
    st.subheader("Tabla: Fuentes del capital a necesitar")
    col1, col2 = st.columns(2)
    with col1:
        a = st.number_input("Mi propio dinero (Lps)", 0, 10000000000000, 50000)
        b = st.number_input("Ayuda de mi familia y remesa", 0, 1000000000000, 30000)
        c = st.number_input("Otros fuentes del fondo propio", 0, 1000000000000, 0)
        st.write(f"<p style='text-align: right;'>Monto total de capital propio: {a+b+c} Lps.</p>", unsafe_allow_html=True)

    with col2:
        d = st.number_input("Crédito de la cooperativa", 0, 1000000000000000, 100000)
        e = st.number_input("Otros créditos", 0, 1000000000000000, 0)
        st.write(f"<p style='text-align: right;'>Monto total de créditos: {d+e} Lps.</p>", unsafe_allow_html=True)
        st.write(f"<p style='text-align: right;'>Total: {a+b+c+d+e} Lps.</p>", unsafe_allow_html=True)

    # 合計の比較
    total3= a+b+c+d+e
    if total1 + total2 > total3:
        st.warning(f"¡Ojo! La financiación (fuentes del capital) es insuficiente para el monto total necesario para montar el negocio. La cantidad necesaria debe ser {total1 + total2} Lps.")
    elif total1 + total2 < total3:
        st.warning(f"¡Ojo! La financiación (fuentes del capital) excede el monto total necesario para montar el negocio. La cantidad necesaria debe ser {total1 + total2} Lps.")

    # エクセル出力のための関数（editable_df1 と editable_df2 を使用）
    def to_excel(editable_df1, editable_df2, a, b, c, d, e):
        output = BytesIO()
        workbook = Workbook()
        
        # シート1: Editableなテーブル1
        worksheet1 = workbook.active
        worksheet1.title = "Tabla 1"
        for r in dataframe_to_rows(editable_df1, index=False, header=True):
            worksheet1.append(r)
        
        # シート2: Editableなテーブル2
        worksheet2 = workbook.create_sheet("Tabla 2")
        for r in dataframe_to_rows(editable_df2, index=False, header=True):
            worksheet2.append(r)
        
        # シート3: Capital sources table
        worksheet3 = workbook.create_sheet("Tabla 3")
        worksheet3.append(["Asuntos", "Monto (Lps.)"])
        worksheet3.append(["Mi propio dinero", a])
        worksheet3.append(["Ayuda de mi familia y remesa", b])
        worksheet3.append(["Otros fuentes del fondo propio", c])
        worksheet3.append(["Monto total de capital propio", a+b+c])
        worksheet3.append(["Crédito de la cooperativa", d])
        worksheet3.append(["Otros créditos", e])
        worksheet3.append(["Monto total de créditos", d+e])
        worksheet3.append(["Total", a+b+c+d+e])
        
        # 各シートのカラム幅を設定
        for ws in workbook.worksheets:
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30

        workbook.save(output)
        return output.getvalue()

    # エクセル出力のためのボタン (editable_df1 と editable_df2 を使用)
    excel_data = to_excel(editable_df1, editable_df2, a, b, c, d, e)
    st.download_button(label="Descargar en Excel", data=excel_data, file_name="planificacion_de_capital.xlsx")
